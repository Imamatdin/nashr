"""XLSX parser built on openpyxl.

Each worksheet becomes one :class:`ParsedPage` whose text is a tab/newline
serialization of every populated cell, with the same content also captured as
a structured table in ``page.tables``. XLSX files never need OCR.
"""

from __future__ import annotations

import asyncio
import logging
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from packages.core.models.source import (
    ParsedPage,
    ParsedSource,
    SourceMetadataExtracted,
)
from packages.workers.source.parsers.lang_detect import detect_language

logger = logging.getLogger(__name__)


class XLSXParser:
    """Parses an XLSX byte stream into a :class:`ParsedSource`."""

    async def parse(self, file_bytes: bytes, filename: str) -> ParsedSource:
        return await asyncio.to_thread(self._parse_sync, file_bytes, filename)

    def _parse_sync(self, file_bytes: bytes, filename: str) -> ParsedSource:
        errors: list[str] = []

        try:
            workbook = load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
        except Exception as exc:
            logger.warning("Failed to open XLSX %s: %s", filename, exc)
            return ParsedSource(
                filename=filename,
                file_type="xlsx",
                file_size_bytes=len(file_bytes),
                pages=[],
                metadata=SourceMetadataExtracted(),
                full_text="",
                needs_ocr_pages=[],
                parse_errors=[f"Failed to open XLSX: {exc}"],
            )

        pages: list[ParsedPage] = []
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets):
                pages.append(self._parse_sheet(sheet, sheet_index))
        except Exception as exc:
            errors.append(f"Worksheet walk failed: {exc}")
            logger.warning("XLSX worksheet walk failed: %s", exc)
        finally:
            workbook.close()

        if not pages:
            pages = [ParsedPage(page_number=1, text="", char_count=0, needs_ocr=False)]

        full_text = "\n\n".join(p.text for p in pages if p.text)
        word_count = len(full_text.split()) if full_text else 0

        metadata = SourceMetadataExtracted(
            page_count=len(pages),
            word_count=word_count,
            language_detected=detect_language(full_text),
            has_images=False,
        )

        return ParsedSource(
            filename=filename,
            file_type="xlsx",
            file_size_bytes=len(file_bytes),
            pages=pages,
            metadata=metadata,
            full_text=full_text,
            needs_ocr_pages=[],
            parse_errors=errors,
        )

    @staticmethod
    def _parse_sheet(sheet: Worksheet, sheet_index: int) -> ParsedPage:
        rows_text: list[str] = []
        table: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [_render_cell(cell) for cell in row]
            if not any(cell for cell in cells):
                continue
            table.append(cells)
            rows_text.append("\t".join(cells))

        sheet_title = (sheet.title or f"Sheet{sheet_index + 1}").strip()
        sections: list[str] = [sheet_title] if sheet_title else []
        sections.extend(rows_text)
        text = "\n".join(sections)

        return ParsedPage(
            page_number=sheet_index + 1,
            text=text,
            char_count=sum(1 for ch in text if not ch.isspace()),
            needs_ocr=False,
            headings=[sheet_title] if sheet_title else [],
            tables=[table] if table else [],
        )


def _render_cell(value: object) -> str:
    """Stringify one openpyxl cell value without losing zero/blank semantics."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        # Preserve integers without trailing ".0"; floats keep precision.
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()
